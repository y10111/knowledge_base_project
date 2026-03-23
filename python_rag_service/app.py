"""
RAG 服务 - 加速版 + 强制使用检索片段
- 大模型独占MPS/GPU资源
- 向量模型保留在CPU
- 优化生成速度
- 修改提示词，确保有片段时基于片段回答
"""

import os
os.environ["HF_HUB_OFFLINE"] = "1"
os.environ["TRANSFORMERS_OFFLINE"] = "1"
os.environ["HF_DATASETS_OFFLINE"] = "1"
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
os.environ["PYTORCH_ENABLE_MPS_FALLBACK"] = "1"

import json
import numpy as np
import torch
import requests
from flask import Flask, request, jsonify
from transformers import AutoTokenizer, AutoModel
from transformers import AutoTokenizer, AutoModelForCausalLM
import faiss
import pdfplumber
from docx import Document as DocxDocument
from openpyxl import load_workbook

app = Flask(__name__)

# 导入CORS配置（请确保 config.py 存在并包含 setup_cors 函数）
from config import setup_cors
setup_cors(app)

# =====================================================
# 配置参数
# =====================================================
CONFIG = {
    "chunk_size": 500,
    "chunk_overlap": 50,
    "retrieval_top_k": 3,
    "similarity_threshold": 0.5,
    "max_new_tokens": 512,
    "temperature": 0.1,          # 降低温度使输出更确定，减少采样开销
    "do_sample": False,           # 保持贪心解码，最快
}

# =====================================================
# 模型路径
# =====================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EMBED_MODEL_PATH = os.path.join(BASE_DIR, "models", "bge-small-zh-v1.5")
LLM_MODEL_PATH = os.path.join(BASE_DIR, "models", "Qwen2.5-1.5B-Instruct")

print("=" * 50)
print("RAG 服务启动（加速版 + 强制使用片段）")
print("向量模型路径:", EMBED_MODEL_PATH)
print("大模型路径:", LLM_MODEL_PATH)
print("配置:", CONFIG)
print("=" * 50)

# =====================================================
# 加载 BGE 向量模型（CPU）
# =====================================================
print("正在加载 BGE 向量模型（CPU）...")
try:
    bge_tokenizer = AutoTokenizer.from_pretrained(EMBED_MODEL_PATH, trust_remote_code=True, local_files_only=True)
    bge_model = AutoModel.from_pretrained(
        EMBED_MODEL_PATH,
        trust_remote_code=True,
        local_files_only=True,
        device_map="cpu",
        torch_dtype=torch.float32
    ).eval()
    print("✓ BGE 向量模型加载成功")
except Exception as e:
    print("✗ BGE 向量模型加载失败:", e)
    exit(1)

def encode(texts, max_length=512):
    inputs = bge_tokenizer(texts, padding=True, truncation=True, max_length=max_length, return_tensors="pt")
    inputs = {k: v.to("cpu") for k, v in inputs.items()}
    with torch.no_grad():
        outputs = bge_model(**inputs)
    last_hidden = outputs.last_hidden_state
    attention_mask = inputs['attention_mask'].unsqueeze(-1).float()
    sum_embeddings = torch.sum(last_hidden * attention_mask, dim=1)
    sum_mask = torch.clamp(attention_mask.sum(dim=1), min=1e-9)
    mean_embeddings = sum_embeddings / sum_mask
    mean_embeddings = torch.nn.functional.normalize(mean_embeddings, p=2, dim=1)
    return mean_embeddings.numpy()

# =====================================================
# 加载 Qwen 大模型（独占 MPS/GPU）
# =====================================================
print("正在加载 Qwen2.5-1.5B-Instruct 大模型...")
try:
    llm_tokenizer = AutoTokenizer.from_pretrained(LLM_MODEL_PATH, trust_remote_code=True, local_files_only=True)
    if llm_tokenizer.pad_token is None:
        llm_tokenizer.pad_token = llm_tokenizer.eos_token

    # 检查可用的加速设备
    if torch.backends.mps.is_available():
        device = "mps"
        print("使用 MPS (Apple GPU) 加速")
    else:
        device = "cpu"
        print("使用 CPU 运行大模型")

    llm = AutoModelForCausalLM.from_pretrained(
        LLM_MODEL_PATH,
        trust_remote_code=True,
        local_files_only=True,
        device_map=device,                # 强制放在单一设备，避免跨设备通信
        torch_dtype=torch.float16,        # 半精度加速
        low_cpu_mem_usage=True
    ).eval()
    print(f"✓ Qwen 大模型加载成功，使用设备: {llm.device}")
except Exception as e:
    print("✗ Qwen 大模型加载失败:", e)
    exit(1)

# =====================================================
# 向量索引存储
# =====================================================
chunks = []
index = None

# =====================================================
# 从 Java 后端重建索引
# =====================================================
def rebuild_index_from_java(java_base_url="http://localhost:8080"):
    """从 Java 后端获取所有文档并重建索引"""
    try:
        resp = requests.get(f"{java_base_url}/api/docs/all", timeout=30)
        if resp.status_code == 200:
            docs = resp.json()
            print(f"从 Java 获取到 {len(docs)} 个文档，开始重建索引...")
            global chunks, index
            chunks = []
            all_vectors = []
            for doc in docs:
                doc_id = doc['id']
                content = doc['content']
                text_chunks = split_text(content)
                if text_chunks:
                    vectors = encode(text_chunks)
                    for i, chunk in enumerate(text_chunks):
                        chunks.append({'doc_id': doc_id, 'text': chunk})
                        all_vectors.append(vectors[i])
            if all_vectors:
                vectors_array = np.array(all_vectors).astype('float32')
                dimension = vectors_array.shape[1]
                index = faiss.IndexFlatIP(dimension)
                faiss.normalize_L2(vectors_array)
                index.add(vectors_array)
            print(f"索引重建完成，共索引 {len(chunks)} 个片段。")
        else:
            print(f"从 Java 获取文档失败，状态码: {resp.status_code}")
    except Exception as e:
        print(f"重建索引时出错: {e}")

# =====================================================
# 辅助函数
# =====================================================
def split_text(text, chunk_size=None, overlap=None):
    chunk_size = chunk_size or CONFIG["chunk_size"]
    overlap = overlap or CONFIG["chunk_overlap"]
    chunks = []
    start = 0
    while start < len(text):
        end = min(start + chunk_size, len(text))
        chunks.append(text[start:end])
        start += chunk_size - overlap
    return chunks

def extract_text_from_file(file_path, file_type):
    text = ""
    if file_type == "pdf":
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    elif file_type == "docx":
        doc = DocxDocument(file_path)
        for para in doc.paragraphs:
            text += para.text + "\n"
    elif file_type == "txt":
        with open(file_path, "r", encoding="utf-8") as f:
            text = f.read()
    elif file_type == "xlsx":
        wb = load_workbook(file_path, data_only=True)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join(str(cell) for cell in row if cell)
                if row_text:
                    text += row_text + "\n"
    else:
        raise ValueError(f"不支持的文件类型: {file_type}")
    return text.strip()

# =====================================================
# 通用生成函数（优化版）
# =====================================================
def generate_answer(messages, thought_process):
    prompt = llm_tokenizer.apply_chat_template(messages, tokenize=False, add_generation_prompt=True)
    inputs = llm_tokenizer(prompt, return_tensors="pt").to(llm.device)
    with torch.no_grad():  # 确保推理时不计算梯度
        outputs = llm.generate(
            inputs.input_ids,
            attention_mask=inputs.attention_mask,
            max_new_tokens=CONFIG["max_new_tokens"],
            temperature=CONFIG["temperature"],
            do_sample=CONFIG["do_sample"],
            pad_token_id=llm_tokenizer.pad_token_id,
            num_beams=1,                # 束搜索宽度1 = 贪心，最快
            repetition_penalty=1.0,
            early_stopping=False,
            use_cache=True              # 使用KV缓存，加速自回归
        )
    input_len = inputs.input_ids.shape[1]
    generated_tokens = outputs[0][input_len:]
    answer = llm_tokenizer.decode(generated_tokens, skip_special_tokens=True).strip()
    return answer

# =====================================================
# 路由：文档索引
# =====================================================
@app.route('/index', methods=['POST'])
def index_document():
    data = request.json
    doc_id = data.get('doc_id')
    content = data.get('content')
    file_path = data.get('file_path')
    file_type = data.get('file_type')

    if not doc_id:
        return jsonify({'error': '缺少 doc_id'}), 400

    if not content and file_path and file_type:
        try:
            content = extract_text_from_file(file_path, file_type)
        except Exception as e:
            return jsonify({'error': f'文件解析失败: {str(e)}'}), 400

    if not content:
        return jsonify({'error': '未提供文档内容'}), 400

    text_chunks = split_text(content)
    if not text_chunks:
        return jsonify({'status': 'ok', 'chunks': 0})

    vectors = encode(text_chunks)

    global chunks, index
    for i, chunk in enumerate(text_chunks):
        chunks.append({'doc_id': doc_id, 'text': chunk})

    if chunks:
        vectors_array = np.array(vectors).astype('float32')
        dimension = vectors_array.shape[1]
        index = faiss.IndexFlatIP(dimension)
        faiss.normalize_L2(vectors_array)
        index.add(vectors_array)

    return jsonify({'status': 'ok', 'chunks': len(text_chunks)})

# =====================================================
# 路由：删除文档索引
# =====================================================
@app.route('/delete', methods=['POST'])
def delete_document():
    data = request.json
    doc_id = data.get('doc_id')
    if not doc_id:
        return jsonify({'error': '缺少 doc_id'}), 400

    global chunks, index

    before_count = len(chunks)
    chunks = [c for c in chunks if c['doc_id'] != doc_id]
    removed = before_count - len(chunks)

    if chunks:
        vectors = encode([c['text'] for c in chunks])
        vectors_array = np.array(vectors).astype('float32')
        dimension = vectors_array.shape[1]
        index = faiss.IndexFlatIP(dimension)
        faiss.normalize_L2(vectors_array)
        index.add(vectors_array)
    else:
        index = None

    return jsonify({'status': 'ok', 'removed': removed, 'remaining': len(chunks)})

# =====================================================
# 路由：智能问答（带回溯过程）
# =====================================================
@app.route('/ask', methods=['POST'])
def ask_question():
    data = request.json
    question = data.get('question')
    if not question:
        return jsonify({'error': '问题不能为空'}), 400

    # 初始化思考过程
    thought = []
    thought.append(f"收到问题：{question}")

    # 无索引情况
    if index is None or index.ntotal == 0:
        thought.append("知识库为空，无法检索相关文档。")
        thought.append("将直接使用大模型自身知识回答。")
        messages = [
            {"role": "system", "content": "你是一个专业的技术专家。用户的问题在知识库中没有找到相关文档，请根据你自己的知识回答，并先说明未找到相关文档。"},
            {"role": "user", "content": question}
        ]
        answer = generate_answer(messages, thought)
        return jsonify({
            'answer': answer,
            'sources': [],
            'thought_process': "\n".join(thought)
        })

    # 1. 向量检索
    thought.append("正在将问题向量化...")
    q_vec = encode([question]).astype('float32')
    faiss.normalize_L2(q_vec)

    thought.append(f"正在检索相似文档片段（top_k={CONFIG['retrieval_top_k']}）...")
    k = min(CONFIG["retrieval_top_k"], index.ntotal)
    scores, indices = index.search(q_vec, k)

    # 收集相关片段
    relevant_chunks = []
    sources = []
    thought.append(f"检索到 {len(indices[0])} 个候选片段，开始筛选（阈值≥{CONFIG['similarity_threshold']}）")
    for i, idx in enumerate(indices[0]):
        if idx < 0 or idx >= len(chunks):
            continue
        score = scores[0][i]
        if score < CONFIG["similarity_threshold"]:
            thought.append(f"片段 {i+1} 相似度 {score:.3f} 低于阈值，舍弃")
            continue
        chunk = chunks[idx]
        relevant_chunks.append(chunk['text'])
        sources.append({'doc_id': chunk['doc_id'], 'score': float(score)})
        thought.append(f"片段 {i+1} 相似度 {score:.3f}，来自文档 {chunk['doc_id']}")

    # 2. 根据检索结果构造 prompt（关键修改点）
    if relevant_chunks:
        thought.append(f"最终保留 {len(relevant_chunks)} 个相关片段，用于生成上下文。")
        context = "\n\n".join([f"[文档片段{i+1}]: {chunk}" for i, chunk in enumerate(relevant_chunks)])
        # 修改系统消息：强制基于片段回答，移除“如果没有则说没有”的模糊指令
        system_msg = "你是一个专业的技术专家。请严格基于提供的文档片段回答用户的问题。务必从片段中提取信息，如果片段中有相关内容，请直接引用或总结。只有在片段中完全没有相关信息时，才可以说'文档中没有找到相关内容'。"
        # 修改用户消息：将上下文放在前面，问题放在后面，更强调基于片段
        user_msg = f"以下是相关的文档片段：\n{context}\n\n请基于以上片段回答用户的问题：{question}"
        thought.append("正在根据上下文生成答案...")
    else:
        thought.append("未找到足够相似的相关文档片段，将直接使用大模型自身知识回答，并告知用户未找到相关文档。")
        system_msg = "你是一个专业的技术专家。用户的问题在知识库中没有找到相关文档，请根据你自己的知识回答，并先说明未找到相关文档。"
        user_msg = f"问题：{question}"

    messages = [
        {"role": "system", "content": system_msg},
        {"role": "user", "content": user_msg}
    ]

    # 3. 生成答案
    try:
        answer = generate_answer(messages, thought)
        thought.append("答案生成完成。")
    except Exception as e:
        thought.append(f"生成失败: {str(e)}")
        answer = f"生成失败: {str(e)}"
        sources = []
    finally:
        # 清空MPS缓存（如果使用MPS）
        if torch.backends.mps.is_available():
            torch.mps.empty_cache()

    return jsonify({
        'answer': answer,
        'sources': sources,
        'thought_process': "\n".join(thought)
    })

# =====================================================
# 健康检查
# =====================================================
@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'index_size': index.ntotal if index else 0})

# =====================================================
# 启动
# =====================================================
if __name__ == '__main__':
    # 从 Java 后端重建索引
    rebuild_index_from_java()
    # 启动服务
    app.run(host='0.0.0.0', port=5001, debug=False)